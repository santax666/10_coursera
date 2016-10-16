import requests
import random
import string
import argparse
import datetime
from openpyxl import Workbook
from bs4 import BeautifulSoup


def send_get_request(url, payload=None):
    site_data = requests.get(url, params=payload, allow_redirects=False)
    return site_data


def get_all_courses():
    courses_list = []
    url = 'https://api.coursera.org/api/courses.v1'
    fields_val = 'primaryLanguages,subtitleLanguages,startDate,workload'
    payload = {'fields': fields_val}
    resp_json = send_get_request(url, payload).json()
    courses_list = resp_json['elements'].copy()
    next_page = resp_json.get('paging', {}).get('next')
    while next_page is not None:
        payload = {'fields': fields_val, 'start': next_page}
        resp_json = send_get_request(url, payload).json()
        courses_list.extend(resp_json['elements'].copy())
        next_page = resp_json.get('paging', {}).get('next')
    return courses_list


def get_random_courses(courses, quantity_courses=20):
    return random.sample(courses, quantity_courses)


def convert_list_to_str(data):
    return ','.join(data)


def get_url_course(slug):
    course_type = ('learn', 'course',)
    for course in course_type:
        url = 'https://www.coursera.org/{0}/{1}'.format(course, slug)
        http_code = send_get_request(url).status_code
        headers = send_get_request(url).history
        if http_code == 200:
            break
    return url


def convert_posix_to_datetime(value):
    milliseconds_in_second = 1000
    value = int(value) / milliseconds_in_second
    return datetime.datetime.fromtimestamp(value).strftime('%d-%m-%Y')


def get_courses_info(courses):
    courses_list = []
    for course in get_random_courses(courses):
        name = course.get('name')
        primary_languages = course.get('primaryLanguages')
        primary_languages = convert_list_to_str(primary_languages)
        subtitle_languages = course.get('subtitleLanguages')
        subtitle_languages = convert_list_to_str(subtitle_languages)
        start_date = course.get('startDate')
        if start_date is not None:
            start_date = convert_posix_to_datetime(start_date)
        work_load = course.get('workload')
        url = get_url_course(course.get('slug'))
        if url is not None:
            rating = get_rating_course(url)
        else:
            rating = None
        course_info = (name, primary_languages, subtitle_languages,
                       start_date, work_load, rating,)
        courses_list.append(course_info)
    return courses_list


def get_rating_course(url):
    responce = send_get_request(url).content
    course_soup = BeautifulSoup(responce, "html.parser")
    rating = course_soup.find('div', {'class': 'ratings-text bt3-visible-xs'})
    if rating is not None:
        return rating.string


def output_courses_info_to_xlsx(courses_list, filepath):
    work_book = Workbook()
    work_sheet = work_book.active

    sheet_title = ('Название курса', 'Основные языки', 'Дополнительные языки',
                   'Дата начала', ' Количество недель', 'Рейтинг')
    for title_count, title in enumerate(sheet_title, 1):
        work_sheet.cell(row=1, column=title_count).value = title
    for course_number, course in enumerate(courses_list, 2):
        for param_number, param in enumerate(course, 1):
            work_sheet.cell(row=course_number,
                            column=param_number).value = param
    work_book.save(filepath)


def create_parser():
    parser = argparse.ArgumentParser()
    parser.add_argument('-f', '--filepath', default='courses.xlsx',
                        help="XLSX-файл для сохранения")
    return parser


if __name__ == '__main__':
    parser = create_parser()
    namespace = parser.parse_args()
    xlsx_file = namespace.filepath

    courses_list = get_all_courses()
    info_data = get_courses_info(courses_list)
    output_courses_info_to_xlsx(info_data, xlsx_file)
